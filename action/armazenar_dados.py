import os
from openpyxl import Workbook, load_workbook


def armazena_dados(temperatura, umidade, data, erro=None):
    arquivo_planilha = "capturas.xlsx" #define uma variavel com o nome do arquvivo que armazena as capturas
    pasta_armazenamento = "dados_capturados/" #define uma pasta para armazenar o arquivo com as capturas
    caminho = pasta_armazenamento + arquivo_planilha #concatena as variaveis para formar o caminho completo do nome do arquivo
    if not os.path.exists(caminho): #verifica se o arquivo nao existe, criando o arquvio e gravando os cabecalhos
        #cria um novo workbook e coloca os cabecalhos
        wb = Workbook()
        ws = wb.active
        ws.title = 'Capturas'
        ws.append(("Data - Hora","Temperatura", "Umidade"))

        wb.save(caminho)

    if erro:
        return "Favor verificar o coletor de dados!" #quando existir um erro, pede para o usuario verificar o coletor de dados.
    else:
        wb = load_workbook(caminho) #carrega a planilha ja criada
        ws = wb.active #ativa a planilha

        ws.append((data, temperatura, umidade)) #adiciona os dados capturados

        wb.save(caminho) #salvo os dados no arquivo de planilha

        return "Dados salvos com sucesso!" #retorna uma mensagem de que os dados foram salvos com sucesso.